#-*-coding:gb2312-*-
import openpyxl
from xpinyin import Pinyin

def remove_extra_spaces(string):
    # ���ַ������ո�ָ���б�
    words = string.split()
    
    # ���ֻ��һ������,ֱ�ӷ���
    if len(words) == 1:
        return string
    
    # ����һ��������ʣ�ಿ��ƴ��,ȥ��ʣ�ಿ���еĿո�
    result = words[0] + ' ' + ''.join(words[1:])
    return result

def capitalize_string(string):
    # ʹ��title()������ÿ�����ʵ�����ĸ��ɴ�д
    capitalized_string = string.title()
    return capitalized_string

# ��Excel�ļ�
workbook = openpyxl.load_workbook('name.xlsx')
sheet = workbook.active

# ���������б�
for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value
    
    p = Pinyin()
    # ������ת��Ϊƴ��
    pinyin_name = p.get_pinyin(name, ' ')
    # ɾ������ո�
    pinyin_name =  remove_extra_spaces(pinyin_name)
    # ����ĸ��д
    pinyin_name = capitalize_string(pinyin_name)
    
    # ��ƴ��д��Excel�ļ�
    sheet.cell(row=row, column=2).value = pinyin_name

# ����Excel�ļ�
workbook.save('names_with_pinyin.xlsx')